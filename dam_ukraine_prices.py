import requests
from bs4 import BeautifulSoup
from io import StringIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

url = "https://www.oree.com.ua/index.php/pricectr/data_view"
headers = {
    "User-Agent": "Mozilla/5.0",
    "Referer": "https://www.oree.com.ua/index.php/pricectr",
    "X-Requested-With": "XMLHttpRequest",
}

payload = {"date": "04.2026", "market": "DAM", "zone": "IPS"}
filename = f"oree_prices_{payload['market']}_{payload['zone']}_{payload['date'].replace('.', '_')}.xlsx"

resp = requests.post(url, data=payload, headers=headers)
resp.encoding = "utf-8"

content = resp.json().get("content", "")
soup = BeautifulSoup(content, "html.parser")
table = soup.find("table")
df = pd.read_html(StringIO(str(table)))[0]

# Reshape to long format: Date | Hour | Price
df_long = df.melt(id_vars=df.columns[0], var_name="Hour", value_name="Price (грн/МВт.год)")
df_long.rename(columns={df.columns[0]: "Date"}, inplace=True)
df_long = df_long[["Date", "Hour", "Price (грн/МВт.год)"]]

# Sort by Date first, then Hour
df_long["Hour"] = pd.to_numeric(df_long["Hour"], errors="coerce")
df_long = df_long.sort_values(["Date", "Hour"]).reset_index(drop=True)

# Build Excel
wb = Workbook()
ws = wb.active
ws.title = "Prices"

header_font = Font(bold=True, name="Arial", color="FFFFFF")
header_fill = PatternFill("solid", start_color="4472C4", end_color="4472C4")
header_align = Alignment(horizontal="center")

# Write headers
for col_idx, col_name in enumerate(df_long.columns, start=1):
    cell = ws.cell(row=1, column=col_idx, value=str(col_name))
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

# Write data
for row_idx, row in df_long.iterrows():
    for col_idx, value in enumerate(row, start=1):
        ws.cell(row=row_idx + 2, column=col_idx, value=value)

# Auto-fit column widths
for col in ws.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    ws.column_dimensions[col[0].column_letter].width = max_len + 4

wb.save(filename)
from google.colab import files
files.download(filename)
